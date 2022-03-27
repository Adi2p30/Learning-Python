import webbrowser
import discord
import os
answer1 = ""
client = discord.Client()
import random
import datetime
import math
import openpyxl

diarydate = {}
from openpyxl import load_workbook

now = datetime.datetime.now()
ppl = {}


@client.event
async def on_ready():
    print("logged in as {0.user}".format(client))


@client.event
async def on_message(message):
    def chatbot():
        chat = ""
        hellow = ["hi", "is anyone there?", "hello", "good day", "hello there", "hey"]
        bye = ["cya", "see you later", "bye", "goodbye", "i am Leaving", "have a good day", "bye bye", "byee",
               "cya later", "peace out"]
        howare = ["how are you", "whats up", "how you doing", "wassup", "watcha doing"]
        name = ["whats your name", "what is your name", "do you have any name", "what should i call you", ]
        bad = ["bad", "sad", "not great", "not good", "very sad", "very bad"]
        good = ["great", "good", "amazing", "tremendous", "spectacular"]
        if chat.lower() in hellow:
            botans = ["Hello !", "hi", "hi there", "hey"]
            answer1 = (random.choice(botans))

        elif chat.lower() in howare:
            botans = ["I am fine How are you? ", "thanks How are you? ", "Happy How are you?", "I am good How are you?"]
            answer1 = (random.choice(botans))
        elif chat.lower() in good:
            botans = ["Great !", "Thats amazing", "Same!", ":)"]
            answer1 = (random.choice(botans))
        elif chat.lower() in bad:
            botans = ["well maybe i can cheer you up!", "oh thats sad hope you have a great day", ":(",
                      "well you will be fine"]
            answer1 = (random.choice(botans))

        elif chat.lower() in name:
            botans = ["My name is THE BOT", "You can call me THE BOT", "THE BOT in your service",
                      "My friends call me THE BOT"]
            answer1 = (random.choice(botans))

        elif chat.lower() in bye:
            botans = ["Sad to see you go :(", "Talk to you later", "Goodbye!", "Have a nice Day"]
            answer1 = (random.choice(botans))
        else:
            answer1 = ("Sorry What ?")

        def store():

            workbook = load_workbook(filename='Mybotdata.xlsx')
            sheet = workbook.active
            name = input("Name: ")
            if name in ppl:
                answer1 = ("This person has already been added")
            else:
                DOB = input("DOB (add spaces between numbers): ")
                year = (DOB[6] + DOB[7] + DOB[8] + DOB[9])
                findage = (int(now.year) - int(year))
                School = input("School/Job: ")
                Relation = input("Relation (related to you): ")
                ppl[name] = (name, DOB, School, Relation)
            row = sheet.max_row + 1
            for key, values in ppl.items():
                sheet.cell(row=row, column=1, value=key)
                column = 2
                for element in values:
                    sheet.cell(row=row, column=column, value=element)
                    column += 1
                row += 1
            workbook.save(filename="Mybotdata.xlsx")
            answer1 = ("data stored")

    def info():
        name = input("Name: ")
        wb = load_workbook(filename='Mybotdata.xlsx')
        sheet_ranges = wb.active
        for i in range(1, sheet_ranges.max_row + 1):
            if sheet_ranges.cell(row=i, column=1).value == name:
                name = sheet_ranges.cell(row=i, column=2).value
                DOB = sheet_ranges.cell(row=i, column=3).value
                year = (DOB[6] + DOB[7] + DOB[8] + DOB[9])
                findage = (int(now.year) - int(year))
                School = sheet_ranges.cell(row=i, column=4).value
                Relation = sheet_ranges.cell(row=i, column=5).value
                answer1 = (("Name: " + str(name) + "\n"
                            "DOB: " + str(DOB) + "\n"
                            "age: " + str(findage) + "\n"
                            "School/Job: " + str(School) + "\n"
                            "Relation: " + str(Relation)))

    def delete():
        wb = load_workbook(filename='Mybotdata.xlsx')
        sheet_ranges = wb.active
        namedel = input("who do you want to delete? \n")
        for i in range(1, sheet_ranges.max_row + 1):
            if sheet_ranges.cell(row=i, column=1).value == namedel:
                sheet_ranges.delete_rows(i)
                wb.save(filename="Mybotdata.xlsx")
                answer1 = (namedel + " was deleted")

    def Diary():
        import datetime
        now = datetime.datetime.now()
        y = 1
        didtoday = str(now.date()) + "\n"
        answer1 = (didtoday)
        x = input("What did you do today \n")
        didtoday = didtoday + "1" + "." + x
        answer1 = (didtoday)
        while x != "done":
            y = y + 1
            x = input("\n What did you do today \n")
            if x != "done":
                didtoday = didtoday + "\n" + str(y) + "." + x
                answer1 = (didtoday)
            else:
                break
        diarydate[str(now.date())] = didtoday

    def Ngp():
        d = int(input("enter increment ratio: "))
        a = int(input("eneter start number (non 0): "))
        f = int(input("enter numbers of terms: "))
        sum = a
        for a in range(a, f):
            sum = sum * d
            answer1 = (sum)
        answer1 = ("_______________________")
        answer1 = ("your nth term is:" + sum)

    def Sap():
        a = int(input("what is the start number? "))
        t = int(input("what is the ending number "))
        d = int(input("what is the difference: "))
        sum = a
        for a in range(a, t):
            a = a + d
            sum = sum + a
        answer1 = ("______________________________________________________")
        answer1 = ("Your answer is")
        answer1 = (sum)
        answer1 = ("______________________________________________________")
        answer1 = ("______________________________________________________")

    def Saltpsq():
        a = int(input("what is the start number? "))
        t = int(input("what is the ending number "))
        sum = a * a
        for a in range(a, t):
            a = a + 1
            if (a % 2) == 1:
                a = a
                sum = sum - (a * a)
                answer1 = (sum)
            else:
                a = abs(a)
                sum = sum + (a * a)
                answer1 = (sum)
        answer1 = ("Your answer is " + (sum))

    def Saltpcb():
        a = int(input("what is the start number? "))
        t = int(input("what is the ending number "))
        sum = a * a
        for a in range(a, t):
            a = a + 1
            if (a % 2) == 1:
                a = a
                sum = sum - (a * a * a)
                answer1 = (sum)
            else:
                a = abs(a)
                sum = sum + (a * a * a)
                answer1 = (sum)
        answer1 = ("Your answer is " ,(sum))


    def Nap():
        d = int(input("enter increment number: "))
        a = int(input("eneter start number: "))
        f = int(input("enter numbers of terms: "))
        sum = 0
        for a in range(a, f):
            sum = sum + d
        answer1 = ("your nth term is: " + str(sum))

    def QE():
        a = int(input("a"))
        b = int(input("b"))
        c = int(input("c"))
        delta = (math.sqrt((b * b) - (4 * a * c)))
        root = (-b + delta) / (2 * a)
        root1 = (-b - delta) / (2 * a)
        vertexx = -b / (2 * a)
        vertexy = -(delta * delta) / (4 * a)
        answer1 = ("roots = " + str(root) + ", " + str(root1))
        ("\nvertex = " + "(" + str(vertexx) + ", " + str(vertexy) + ")")
        ("\ndelta = " + str(delta * delta))
        ("\nSigma1 is " + str((-b / a)) + "\n" + "Sigma2 is " + str((c / a)))

    def PT():
        a = int(input())
        b = int(input())
        answer1 = (math.sqrt(a * a + b * b))

    def Slope():
        x = int(input("x1"))
        x1 = int(input("x2"))
        y = int(input("y1"))
        y1 = int(input("y2"))
        answer1 = ((y1 - y) / (x1 - x))

    def sqcbchart():
        count1 = 0
        count2 = 0
        answer1 = ("Number" + " " + "Sqaures" + " " + "Cubes")
        for i in range(0, 100):
            answer1 = (str(i) + (" " * 6) + str((i ** 2)) + (" " * 7) + str((i ** 3)))

    def distance():
        x1 = int(input("X1: "))
        x2 = int(input("X2: "))
        y1 = int(input("Y1: "))
        y2 = int(input("Y2: "))
        withoutsqrt = ((x2 - x1) ** 2) + ((y2 - y1) ** 2)
        answer1 = ("without swrt: ", withoutsqrt,"\n",str(math.sqrt(withoutsqrt)))

    def Sapsq():
        n = int(input("what end number"))
        sum = ((n * (n + 1) * (2 * n + 1)) / 6)
        answer1 = ("Your answer is",int(sum))

    def links():
        answer1 = (
            " Canva \n Youtube \n Socials \n Netflix \n Amazon Prime \n Gmail \n Instagram \n School \n Maths \n Physics \n Chem")
        link = input("Which Links would you like to open?\n")
        if link == "Canva":
            webbrowser.open('https://www.canva.com/')
        if link == "Youtube":
            webbrowser.open('https://www.youtube.com/')
        if link == "Socials":
            webbrowser.open('https://mail.google.com/mail/u/0/#inbox')
            webbrowser.open('instagram.com/?hl=en')
            webbrowser.open('https://twitter.com/home?lang=en')
        if link == "Netflix":
            webbrowser.open('https://www.netflix.com/browse')
        if link == "Amazon Prime":
            webbrowser.open('https://www.primevideo.com/?ref_=dvm_pds_amz_in_as_s_g_313art|m_woIqHOSOc_c442183707292')
        if link == "Gmail":
            webbrowser.open('https://mail.google.com/mail/u/0/#inbox')
        if link == "Instagram":
            webbrowser.open('instagram.com/?hl=en')
        if link == "School":
            webbrowser.open(
                'https://meetingsapac27.webex.com/webappng/sites/meetingsapac27/meeting/download/d0e1941446bd9a64471ab29c58599b98')
        if link == "Maths":
            webbrowser.open('https://us02web.zoom.us/j/87084998112?pwd=c2JQV3JGWHJBcmZ2dzFkemNtU3RYUT09')
        if link == "Physics":
            webbrowser.open('https://us02web.zoom.us/j/5500115066?pwd=Q29MWnZrWFFEMnp5bmd4RGNobUFIQT09')
        if link == "Chem":
            webbrowser.open('')
        else:
            answer1 = ("Link doesnt exist in database!")
    functions = ['Sapsq', 'Slope', 'Nap', 'distance', 'store', 'info', 'delete', 'paary', 'links', 'Diary',
                     'emotionsave', 'google', 'youtube', 'music', 'search', 'help2']
    def help2():
        functions = ['Sapsq', 'Slope', 'Nap', 'distance', 'store', 'info', 'delete', 'paary', 'links', 'Diary',
                     'emotionsave', 'google', 'youtube', 'music', 'search', 'help2']
        message.channel.send(functions)

    if message.author == client.user:
       return
    if message.content in functions:
         eval(message.content + "()")
    else:
         chatbot()
    await message.channel.send(answer1)


client.run(os.getenv("Token"))